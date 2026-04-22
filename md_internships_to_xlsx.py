#!/usr/bin/env python3
"""Build internships.xlsx from markdown clippings in 'MD files/'."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

HEADERS = [
    "Company",
    "Title",
    "Category",
    "Location",
    "Title_URL",
    "Status",
    "Date",
    "About_the_job",
    "Posted_time",
    "Salary",
    "People_applied",
    "Company_URL",
    "Company_follower",
    "Company_size",
    "Count_of_employee_onLinkedIn",
    "Company_Intro",
    "Is this an internship",
]

MD_DIR = Path(__file__).resolve().parent / "MD files"
OUTPUT = Path(__file__).resolve().parent / "internships_from_md.xlsx"


def parse_frontmatter(raw: str) -> tuple[dict[str, str], str]:
    if not raw.startswith("---"):
        return {}, raw
    end = raw.find("\n---", 3)
    if end == -1:
        return {}, raw
    block = raw[3:end].strip("\n")
    body = raw[end + 4 :].lstrip("\n")
    meta: dict[str, str] = {}
    for line in block.split("\n"):
        if ":" not in line:
            continue
        k, _, v = line.partition(":")
        k, v = k.strip(), v.strip()
        if v.startswith('"') and v.endswith('"'):
            v = v[1:-1]
        meta[k] = v
    return meta, body


def split_title_company(title_field: str) -> tuple[str, str]:
    if " | " in title_field:
        left, right = title_field.rsplit(" | ", 1)
        return left.strip(), right.strip()
    return title_field.strip(), ""


def extract_section(body: str, header: str) -> str | None:
    pat = re.escape(header) + r"\s*\n+"
    m = re.search(pat, body)
    if not m:
        return None
    start = m.end()
    rest = body[start:]
    nxt = re.search(r"\n### [^\n]+\n", rest)
    if nxt:
        return rest[: nxt.start()].strip()
    nxt2 = re.search(r"\n## [^\n]+\n", rest)
    if nxt2:
        return rest[: nxt2.start()].strip()
    return rest.strip()


def parse_handshake(body: str, meta: dict[str, str]) -> dict[str, str]:
    row = {h: "" for h in HEADERS}
    title, company = split_title_company(meta.get("title", ""))
    row["Title"] = title
    row["Company"] = company
    row["Title_URL"] = meta.get("source", "")
    row["Date"] = meta.get("created", "")

    posted = re.search(r"Posted\s+(.+?)(?=Apply by|\Z)", body)
    if posted:
        row["Posted_time"] = posted.group(1).strip()

    glance = extract_section(body, "### At a glance")
    if glance:
        lines = [ln.strip() for ln in glance.split("\n") if ln.strip()]
        salary_bits = []
        loc_bits = []
        for ln in lines:
            if re.search(r"\$|/hr|Unpaid|paid internship", ln, re.I):
                salary_bits.append(ln)
            if re.search(
                r"based in|Remote|On-?site|Hybrid|In person|from one of|from the location",
                ln,
                re.I,
            ):
                loc_bits.append(ln)
        if salary_bits:
            row["Salary"] = " ".join(salary_bits)
        if loc_bits:
            row["Location"] = " ".join(loc_bits)

    job = extract_section(body, "### Job description")
    if job:
        row["About_the_job"] = job

    emp_m = re.search(
        r"### About the employer\s+(.*?)(?=\n### [^\n]+\n|\Z)",
        body,
        re.DOTALL,
    )
    if emp_m:
        emp = emp_m.group(1)
        emp = re.sub(r"!\[[^\]]*\]\([^)]*\)\s*", "", emp)
        co = re.search(r"####\s*\[([^\]]+)\]\(([^)]+)\)", emp)
        if co:
            row["Company_URL"] = co.group(2)
            if not row["Company"]:
                row["Company"] = co.group(1)
        for ln in emp.split("\n"):
            ln = ln.strip()
            if re.search(
                r"\d[\d,]*\s*-\s*[\d,]+\s+employees|\d[\d,]*\+?\s*employees|employees worldwide",
                ln,
                re.I,
            ) and "employee" in ln.lower():
                row["Company_size"] = ln
                break
        after_learn = re.split(r"\[Learn more about[^\]]*\]", emp, 1)[0]
        intro = re.sub(r"####\s*\[[^\]]+\]\([^)]+\)\s*", "", after_learn)
        intro = re.sub(r"!\[[^\]]*\]\([^)]*\)\s*", "", intro)
        intro = intro.strip()
        if intro:
            row["Company_Intro"] = re.sub(r"\s+", " ", intro)[:8000]

    row["Is this an internship"] = "Yes"
    return row


def parse_linkedin(body: str, meta: dict[str, str]) -> dict[str, str]:
    row = {h: "" for h in HEADERS}
    title, company = split_title_company(meta.get("title", ""))
    row["Title"] = title
    row["Company"] = company
    row["Title_URL"] = meta.get("source", "")

    row["Date"] = meta.get("created", "")

    first_co = re.search(r"^\[([^\]]+)\]\((https://www\.linkedin\.com/company/[^)]+)\)", body, re.M)
    if first_co:
        if not row["Company"]:
            row["Company"] = first_co.group(1)
        row["Company_URL"] = first_co.group(2)

    meta_line = re.search(
        r"^([^\n]+?)\s*·\s*Reposted\s+([^\n·]+)\s*·\s*(.+)$",
        body,
        re.M,
    )
    if meta_line:
        row["Location"] = meta_line.group(1).strip()
        row["Posted_time"] = "Reposted " + meta_line.group(2).strip()
        row["People_applied"] = meta_line.group(3).strip()

    job_m = re.search(
        r"## About the job\s*\n+(.*?)(?=\n## |\Z)",
        body,
        re.DOTALL,
    )
    if job_m:
        row["About_the_job"] = job_m.group(1).strip()

    aj = row["About_the_job"] or ""
    clean = re.sub(r"\*+", "", aj)
    clean = re.sub(r"\\([\-–])", r"\1", clean)
    sal = re.search(
        r"\$[\d,]+\s*[-–]\s*\$?[\d,]+(?:\s*(?:per\s*)?hour)?",
        clean,
        re.I,
    )
    if not sal:
        sal = re.search(r"\$[\d,]+(?:\s*/?\s*(?:per\s*)?hour)?", clean, re.I)
    if sal:
        row["Salary"] = sal.group(0).strip()

    co_m = re.search(
        r"## About the company\s*\n+(.*?)(?=\n## |\Z)",
        body,
        re.DOTALL,
    )
    if co_m:
        block = co_m.group(1).strip()
        lines = [ln.strip() for ln in block.split("\n") if ln.strip()]
        if lines:
            first = lines[0]
            if "·" not in first and not re.search(r"employees|LinkedIn", first, re.I):
                row["Category"] = first
        em = re.search(r"([\d,]+\s*-\s*[\d,]+\s+employees|[\d,]+\+\s*employees)", block, re.I)
        if em:
            row["Company_size"] = em.group(1)
        li = re.search(r"([\d,]+)\s+on\s+LinkedIn", block, re.I)
        if li:
            row["Count_of_employee_onLinkedIn"] = li.group(1)
        intro_start = 0
        if lines and row["Category"] and lines[0] == row["Category"]:
            intro_start = 1
        while intro_start < len(lines):
            ln = lines[intro_start]
            if ln in ("•",) or re.fullmatch(r"[\d,]+", ln):
                intro_start += 1
                continue
            if re.search(r"employees|on LinkedIn", ln, re.I):
                intro_start += 1
                continue
            break
        intro_lines = lines[intro_start:]
        intro = " ".join(intro_lines)
        intro = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", intro)
        row["Company_Intro"] = re.sub(r"\s+", " ", intro).strip()[:8000]

    row["Is this an internship"] = "Yes"
    return row


def parse_file(path: Path) -> dict[str, str]:
    raw = path.read_text(encoding="utf-8", errors="replace")
    meta, body = parse_frontmatter(raw)
    src = meta.get("source", "")
    if "linkedin.com" in src.lower() or "## About the job" in body:
        return parse_linkedin(body, meta)
    return parse_handshake(body, meta)


def main() -> None:
    paths = sorted(
        p
        for p in MD_DIR.glob("*.md")
        if p.name.lower() not in ("readme.md", "cv.md")
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Internships"
    ws.append(HEADERS)
    for p in paths:
        if not p.is_file():
            continue
        row = parse_file(p)
        ws.append([row.get(h, "") for h in HEADERS])

    for i, _ in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(48, 14 + (5 if i == 8 else 0))

    wb.save(OUTPUT)
    print(f"Wrote {len(paths)} rows to {OUTPUT}")


if __name__ == "__main__":
    main()
